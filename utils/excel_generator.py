from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


def generate_excel(results, output_path):
    """
    Generate Excel report with performance optimizations:
    - Reduced styling operations
    - Limited number of cells with custom styling
    - Batch processing for large result sets
    - Memory usage optimization
    """
    try:
        # Start with a clean workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Domenlarni tekshirish hisoboti"

        # Define headers once
        headers = ["№", "Domen", "Holati", "Holat kodi", "Sahifa turi", "Sarlavha"]

        # Pre-define styles and colors to avoid repeated creation
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        green_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        red_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        left_alignment = Alignment(horizontal="left")

        # Status mappings - define once
        status_codes = {
            200: "OK",
            400: "Tekshirish kerak",
            403: "Tekshirish kerak",
            404: "Topilmadi",
            500: "Server xatosi",
            429: "Tekshirish kerak",
            503: "Tekshirish kerak",
            None: "Mavjud emas"
        }

        page_types = {
            "Internal": "Ichki",
            "External": "Tashqi",
            "Error": "Xato",
            "Non-HTML": "HTML emas",
            "Unknown": "Noma'lum"
        }

        title_defaults = {
            "No Title": "Sarlavhasiz",
            "Error": "Xato",
            "Non-HTML": "HTML emas",
            "Timeout": "Tekshirish kerak"
        }

        # Write headers - just once
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Calculate the number of rows to process for memory efficiency
        total_results = len(results)
        BATCH_SIZE = 1000  # Process 1000 rows at a time to avoid memory issues

        # Process in batches if needed
        for batch_start in range(0, total_results, BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, total_results)
            batch = results[batch_start:batch_end]

            logger.info(f"Processing Excel batch {batch_start + 1}-{batch_end} of {total_results}")

            # Add rows in batches
            for i, result in enumerate(batch, 1):
                row = batch_start + i + 1  # +1 for header row

                # Basic data
                ws.cell(row=row, column=1).value = row - 1
                ws.cell(row=row, column=2).value = result["domain"]

                # Status logic
                status_value = {
                    "Working": "Ishlayapti",
                    "Not Working": "Ishlamayapti",
                    "Need to Check": "Tekshirish kerak"
                }.get(result["status"], "Noma'lum")

                # Special cases for status
                if result["status_code"] in [400, 403]:
                    status_value = "Tekshirish kerak"

                if result.get("title") == "Timeout":
                    status_value = "Tekshirish kerak"

                ws.cell(row=row, column=3).value = status_value

                # Status code
                status_code = result["status_code"]
                status_code_str = status_codes.get(status_code, str(status_code) if status_code else "Mavjud emas")
                ws.cell(row=row, column=4).value = status_code_str

                # Page type and title
                ws.cell(row=row, column=5).value = page_types.get(result["page_type"], result["page_type"])
                ws.cell(row=row, column=6).value = title_defaults.get(result["title"], result["title"])

                # Apply style only to status column for performance
                status_cell = ws.cell(row=row, column=3)

                # Apply appropriate color based on status
                if status_value == "Tekshirish kerak":
                    status_cell.fill = yellow_fill
                elif status_value == "Ishlayapti":
                    status_cell.fill = green_fill
                else:
                    status_cell.fill = red_fill

                # Apply border to all cells in this row (more efficient than iterating again)
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = left_alignment

            # Force garbage collection between batches
            batch = None

        # Set column widths once
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report successfully generated at {output_path}")

        return True
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")

        # Try a minimal report if the full one fails
        try:
            # Create a simpler report with minimal styling
            wb = Workbook()
            ws = wb.active
            ws.title = "Domenlarni tekshirish hisoboti"

            headers = ["№", "Domen", "Holati", "Holat kodi"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col).value = header

            for i, result in enumerate(results[:5000], 2):  # Limit to 5000 rows in emergency
                ws.cell(row=i, column=1).value = i - 1
                ws.cell(row=i, column=2).value = result["domain"]
                ws.cell(row=i, column=3).value = result["status"]
                ws.cell(row=i, column=4).value = result["status_code"]

            wb.save(output_path)
            logger.warning(f"Generated simplified Excel report due to error in main generator")
            return True
        except Exception as backup_error:
            logger.critical(f"Failed to generate even simplified Excel report: {str(backup_error)}")
            return False